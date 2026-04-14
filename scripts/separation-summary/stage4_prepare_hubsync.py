"""
Stage 4 -- Prepare the workbook for the HubSync VBA macro.

Adds the workbook-side scaffolding the HubSync.bas module needs to do
its job when the user imports it into a macro-enabled copy of this
file:

  1. New column N (14) "Synced To Hub" on every FY sheet. HubSync
     writes TODAY() here after a successful PATCH so it never
     double-syncs.
  2. New column O (15) "Do Not Sync" on every FY sheet, with a strict
     Yes/No dropdown. HR can set this to Yes to skip an individual
     row (edge cases, tests, or manual terminations that happened
     outside of the normal workflow).
  3. New "Sync Log" sheet with a header row for the audit trail the
     macro appends to on every run.

Out of scope:
  * The workbook stays as .xlsx. The user imports HubSync.bas into a
    Save-As .xlsm copy via Alt+F11 -> File -> Import File...  The
    install steps are in scripts/separation-summary/README.md and
    mirror the workflow they already use for the Monthly New Hire
    Tracker.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "FY Separation Summary.xlsx"
BACKUP = ROOT / "FY Separation Summary.backup.xlsx"

sys.path.insert(0, str(Path(__file__).parent))
from dv_fixup import DVSpec, apply_dvs  # noqa: E402


FY_SHEETS = [
    "FY 2023 (Jan23-Dec23)",
    "FY 2024 (Jan24-Dec24)",
    "FY 2025 (Jan25-Dec25)",
    "FY 2026 (Jan26-Dec26)",
    "FY 2027 (Jan27-Dec27)",
]

SYNC_LOG_SHEET = "Sync Log"
SYNC_LOG_HEADERS = [
    "Timestamp",
    "Workbook User",
    "FY Sheet",
    "Row",
    "Employee Name",
    "Separation Date",
    "Action",
    "Supabase ID",
    "Match Type",
    "Details",
]


def add_sync_columns(ws) -> None:
    """Add col N (Synced To Hub) + col O (Do Not Sync) to a FY sheet.

    Headers are written into every month's col-header row so they line
    up visually with existing columns."""
    # The 12 col-header rows in the standardized layout are at
    # row 8, 42, 76, 110, 144, 178, 212, 246, 280, 314, 348, 382.
    col_header_rows = [8 + i * 34 for i in range(12)]
    for r in col_header_rows:
        ws.cell(r, 14).value = "Synced To Hub"
        ws.cell(r, 15).value = "Do Not Sync"


def create_sync_log_sheet(wb) -> None:
    if SYNC_LOG_SHEET in wb.sheetnames:
        del wb[SYNC_LOG_SHEET]
    ws = wb.create_sheet(SYNC_LOG_SHEET)
    for c, label in enumerate(SYNC_LOG_HEADERS, start=1):
        cell = ws.cell(1, c)
        cell.value = label
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="left")
    # Give it a reasonable default column width so entries are readable.
    widths = [20, 18, 22, 6, 26, 14, 26, 38, 14, 60]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    # Seed row 2 with a placeholder comment so the sheet isn't
    # surprising when empty. Users can delete it if they want.
    ws.cell(2, 1).value = "(HubSync will append run entries starting here)"
    ws.cell(2, 1).font = Font(italic=True, color="808080")


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"missing {SRC}")

    wb = openpyxl.load_workbook(SRC)

    for name in FY_SHEETS:
        add_sync_columns(wb[name])
    print(f"[stage4] added Synced-To-Hub + Do-Not-Sync columns to {len(FY_SHEETS)} FY sheets")

    create_sync_log_sheet(wb)
    print(f"[stage4] created '{SYNC_LOG_SHEET}' sheet")

    # Add a cell comment on Dashboard explaining how to trigger the
    # macro after they import it.
    dash = wb["Dashboard"]
    dash.cell(5, 1).comment = Comment(
        "After importing HubSync.bas (see scripts/separation-summary/"
        "README.md):\n\n"
        "  * The macro runs automatically on workbook open if auto-sync "
        "is enabled.\n"
        "  * To run manually, Alt+F8 -> HubSync.HubSync -> Run.\n\n"
        "HubSync reads the current FY sheet (this B5 cell) and pushes "
        "any separation whose Date of Separation has arrived to the "
        "hub's Supabase employees table.",
        "Stage 4 setup",
    )

    wb.save(SRC)
    print(f"[stage4] wrote {SRC}")

    # Re-inject data validations, now including the Do Not Sync column.
    strict = lambda last: [
        DVSpec(
            formula="Reference!$B$2:$B$3", sqref=f"E9:E{last}", style="stop",
            error_title="Invalid Rehire",
            error="Rehire Eligibility must be Yes or No.",
        ),
        DVSpec(
            formula="Reference!$G$2:$G$3", sqref=f"F9:F{last}", style="stop",
            error_title="Invalid Status",
            error="Status must be U or D.",
        ),
        DVSpec(
            formula="Reference!$E$2:$E$64", sqref=f"G9:G{last}", style="stop",
            error_title="Invalid Location",
            error="Pick a location from the Reference list.",
        ),
        DVSpec(
            formula="Reference!$C$2:$C$14", sqref=f"H9:H{last}", style="stop",
            error_title="Invalid Reason",
            error="Pick a reason from the Reference list.",
        ),
        DVSpec(
            formula="Reference!$A$2:$A$18", sqref=f"K9:K{last}", style="stop",
            error_title="Invalid Job",
            error="Pick a job from the Reference list.",
        ),
        DVSpec(
            formula="Reference!$L$2:$L$10", sqref=f"M9:M{last}", style="stop",
            error_title="Invalid Department",
            error="Pick a department from the Reference list.",
        ),
        # NEW: Do Not Sync column (col O) -- Yes/No list
        DVSpec(
            formula="Reference!$B$2:$B$3", sqref=f"O9:O{last}", style="stop",
            error_title="Invalid Do Not Sync",
            error="Do Not Sync must be Yes or No (leave blank for No).",
        ),
    ]
    specs = {name: strict(413) for name in FY_SHEETS}
    apply_dvs(SRC, specs)
    print("[stage4] re-applied strict x14 data validations (7 per FY sheet)")


if __name__ == "__main__":
    main()
