"""
Stage 1 -- Reference sheet cleanup.

Scope (Reference sheet only):
  * Trim trailing/leading whitespace on every list entry
  * De-dup within each list, preserving first-seen order
  * Column A (Job Titles): add canonical jobs that appear in real data
  * Column B (Rehire Eligibility): Yes, No -- new, so rehire DV stops being
    a fragile inline CSV string and becomes a range-referenced x14 DV
  * Column C (Reasons for Leaving): add "Resignation without Notice"
  * NEW Column D (Reason Category): Voluntary / Involuntary / Other
    classification for each reason
  * Column G (Status): U, D with a descriptive comment column (H) so users
    remember what U and D mean

Out of scope (later stages):
  * Location column E overhaul  (stage 3 rework)
  * Location -> Department mapping rebuild  (stage 3 rework)
  * Strict validation on FY sheets  (stage 2)

After the openpyxl save we re-inject the x14 data validations for every
FY sheet so we don't lose them. In stage 1 we preserve the original
errorStyle ("warning") -- stage 2 is the one that tightens it to "stop".
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "FY Separation Summary.xlsx"
BACKUP = ROOT / "FY Separation Summary.backup.xlsx"

sys.path.insert(0, str(Path(__file__).parent))
from dv_fixup import DVSpec, apply_dvs  # noqa: E402


# ------------------------------------------------------------------
# Canonical data
# ------------------------------------------------------------------

REASONS: list[tuple[str, str]] = [
    ("Attendance Issues", "Involuntary"),
    ("Better Opportunity", "Voluntary"),
    ("End of Contract", "Other"),
    ("Job Abandonment", "Involuntary"),
    ("Layoff", "Involuntary"),
    ("Performance Issues", "Involuntary"),
    ("Personal Reasons", "Voluntary"),
    ("Relocation", "Voluntary"),
    ("Resignation", "Voluntary"),
    ("Resignation without Notice", "Voluntary"),
    ("Retirement", "Voluntary"),
    ("Termination", "Involuntary"),
    ("Other", "Other"),
]

JOBS: list[str] = [
    "President",
    "VP",
    "Director",
    "DSP",
    "HM",
    "RM",
    "CM",
    "EI",
    "Sub/Floater",
    "Finance",
    "Family Support",
    "Teacher",
    "Assistant Teacher",
    "Job Coach/WI",
    "HR",
    "QA",
    "Other",
]

STATUSES: list[tuple[str, str]] = [
    ("U", "Uneligible for rehire"),
    ("D", "Discharged / terminated"),
]

REHIRE: list[str] = ["Yes", "No"]


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"missing {SRC}")
    if not BACKUP.exists():
        raise SystemExit(f"missing backup at {BACKUP} -- refusing to run")

    wb = openpyxl.load_workbook(SRC)
    ws = wb["Reference"]

    # Clear the columns we own so stale entries don't linger. Leave E
    # (Locations), I/J (Location Mapping), L (Departments) for stage 3.
    for r in range(2, ws.max_row + 1):
        for col in (1, 2, 3, 4, 7, 8):
            ws.cell(r, col).value = None

    # --- header row --------------------------------------------------
    ws.cell(1, 1).value = "Job Titles"
    ws.cell(1, 2).value = "Rehire Eligibility"
    ws.cell(1, 3).value = "Reasons for Leaving"
    ws.cell(1, 4).value = "Reason Category"
    ws.cell(1, 7).value = "Status"
    ws.cell(1, 8).value = "Status Meaning"

    # --- Jobs --------------------------------------------------------
    for i, job in enumerate(JOBS):
        ws.cell(2 + i, 1).value = job

    # --- Rehire ------------------------------------------------------
    for i, val in enumerate(REHIRE):
        ws.cell(2 + i, 2).value = val

    # --- Reasons + category -----------------------------------------
    for i, (reason, cat) in enumerate(REASONS):
        ws.cell(2 + i, 3).value = reason
        ws.cell(2 + i, 4).value = cat

    # --- Status -----------------------------------------------------
    for i, (code, meaning) in enumerate(STATUSES):
        ws.cell(2 + i, 7).value = code
        ws.cell(2 + i, 8).value = meaning

    wb.save(SRC)
    print(f"[stage1] wrote {SRC}")

    # --- Re-inject x14 data validations -----------------------------
    jobs_range = f"Reference!$A$2:$A${1 + len(JOBS)}"
    rehire_range = f"Reference!$B$2:$B${1 + len(REHIRE)}"
    reasons_range = f"Reference!$C$2:$C${1 + len(REASONS)}"
    status_range = f"Reference!$G$2:$G${1 + len(STATUSES)}"
    locations_range = "Reference!$E$2:$E$64"   # untouched in stage 1
    departments_range = "Reference!$L$2:$L$10"  # untouched in stage 1

    def fy_specs(last_row: int) -> list[DVSpec]:
        return [
            DVSpec(
                formula=rehire_range,
                sqref=f"E9:E{last_row}",
                style="warning",
                error_title="Invalid Rehire",
                error="Rehire Eligibility must be Yes or No.",
            ),
            DVSpec(
                formula=status_range,
                sqref=f"F9:F{last_row}",
                style="warning",
                error_title="Invalid Status",
                error="Status must be U (uneligible) or D (discharged).",
            ),
            DVSpec(
                formula=locations_range,
                sqref=f"G9:G{last_row}",
                style="warning",
            ),
            DVSpec(
                formula=reasons_range,
                sqref=f"H9:H{last_row}",
                style="warning",
            ),
            DVSpec(
                formula=jobs_range,
                sqref=f"K9:K{last_row}",
                style="warning",
            ),
            DVSpec(
                formula=departments_range,
                sqref=f"M9:M{last_row}",
                style="warning",
            ),
        ]

    specs = {
        "FY 2023 (Jan23-Dec23)": fy_specs(413),
        "FY 2024 (Jan24-Dec24)": fy_specs(413),
        "FY 2025 (Jan25-Dec25)": fy_specs(413),
        "FY 2026 (Jan26-Dec26)": fy_specs(357),
        "FY 2027 (Jan27-Dec27)": fy_specs(413),
    }
    apply_dvs(SRC, specs)
    print("[stage1] re-injected x14 data validations")


if __name__ == "__main__":
    main()
