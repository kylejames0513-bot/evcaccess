"""
Stage 3 -- Better statistics.

Surgical fixes to the math that drives the Dashboard, without
restructuring the existing layout (it's elaborate and the user has
invested in it -- the right call is to make it correct, not to
redesign it).

Scope:
  1. Centralize the hard-coded "Active Employees" headcount.
     - New table on Reference!N1:O6 keyed by FY year.
     - Each FY sheet's B5 now pulls from that table.
     - Data!B2:B6 unchanged in shape (still points at each FY sheet's
       B5) but now indirectly resolves via Reference.
     This unblocks the Stage 4 VBA macro, which will update the
     current FY row with a live count from Supabase.

  2. Fix the voluntary / involuntary / other classification on the
     Data sheet. Previously Data!F2:H6 used hard-coded COUNTIF chains
     over specific reason strings -- adding "Resignation without
     Notice" in Stage 1 broke that silently. Rewrites to SUMPRODUCT
     over the Data reason table (A55:G67) filtered by the Category
     column, which is itself driven by Reference!D. New reasons added
     to the Reference list are picked up automatically.

  3. Add "Resignation without Notice" to the Data reason table
     (Data!A68) so the count appears in the vol/invol aggregation.

  4. Fix the "Avg Monthly Separations" formula (Data!E2:E6). Used to
     be total / months-with-data, which quietly lied about slow months
     because empty months were excluded from the denominator. Now uses
     months-elapsed (12 for completed years, current month for the
     in-progress year, 0 for future years).

  5. Add two new rows to the Dashboard's "Additional Statistics"
     section:
         C50 = rolling 12-month turnover rate (as of TODAY())
         C51 = avg tenure at separation for the selected FY (FY 2026+
               only, since 2023/2024/2025 have no DOH data)
     A50 / A51 get matching labels. These use SUMPRODUCT across the
     5 FY sheets so the rolling window naturally spans year boundaries.

What this deliberately does NOT touch:
  - The hard-coded EVC Fiscal Year block (Data!B88:M91, A98:I101,
    A111:J114). Those are manual entries spanning two calendar years
    and don't have a clean formulaic fix without a larger rework.
  - The Dashboard's Separations-by-Job and Separations-by-Reason
    tables. They keep using their existing INDEX/MATCH patterns,
    which now resolve correctly because Stage 2 standardized the row
    ranges.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "FY Separation Summary.xlsx"
BACKUP = ROOT / "FY Separation Summary.backup.xlsx"

sys.path.insert(0, str(Path(__file__).parent))
from dv_fixup import DVSpec, apply_dvs  # noqa: E402

# The active-employee counts that were hard-coded in each FY sheet's B5
# on the original workbook. Moving them into one table here.
ACTIVE_EMPLOYEES: list[tuple[str, int]] = [
    ("FY 2023", 339),
    ("FY 2024", 339),
    ("FY 2025", 346),
    ("FY 2026", 366),
    ("FY 2027", 364),
]

FY_SHEETS: list[tuple[str, str]] = [
    ("FY 2023", "FY 2023 (Jan23-Dec23)"),
    ("FY 2024", "FY 2024 (Jan24-Dec24)"),
    ("FY 2025", "FY 2025 (Jan25-Dec25)"),
    ("FY 2026", "FY 2026 (Jan26-Dec26)"),
    ("FY 2027", "FY 2027 (Jan27-Dec27)"),
]


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"missing {SRC}")
    if not BACKUP.exists():
        raise SystemExit(f"missing backup at {BACKUP}")

    wb = openpyxl.load_workbook(SRC)

    # ------------------------------------------------------------------
    # 1. Reference!N:O -- Active Employees by FY, centralized
    # ------------------------------------------------------------------
    ref = wb["Reference"]
    ref.cell(1, 14).value = "FY"                 # N1
    ref.cell(1, 15).value = "Active Emp (FY start)"  # O1
    for i, (fy, count) in enumerate(ACTIVE_EMPLOYEES):
        ref.cell(2 + i, 14).value = fy
        ref.cell(2 + i, 15).value = count
    print(f"[stage3] Reference!N:O populated with {len(ACTIVE_EMPLOYEES)} FY rows")

    # ------------------------------------------------------------------
    # 2. FY sheet B5 cells -> VLOOKUP into the Reference table
    #    (previously hard-coded integers)
    # ------------------------------------------------------------------
    for fy_label, sheet_name in FY_SHEETS:
        ws = wb[sheet_name]
        ws.cell(5, 2).value = (
            f'=IFERROR(VLOOKUP("{fy_label}",Reference!$N$2:$O$6,2,FALSE),0)'
        )
    print("[stage3] FY sheet B5 cells now VLOOKUP Reference!$N$2:$O$6")

    # ------------------------------------------------------------------
    # 3. Data sheet fixes
    # ------------------------------------------------------------------
    d = wb["Data"]

    # 3a. Add "Resignation without Notice" to Data!A68 so its count
    #     appears in the FY columns. Rows 68-74 are empty today so
    #     appending at A68 is safe.
    d.cell(68, 1).value = "Resignation without Notice"
    for i, (_, sheet_name) in enumerate(FY_SHEETS):
        col = 3 + i  # C..G
        d.cell(68, col).value = (
            f"=COUNTIF('{sheet_name}'!H9:H412,\"Resignation without Notice\")"
        )
    print("[stage3] added 'Resignation without Notice' row at Data!A68")

    # 3a.1 Make Data!B56:B68 (the reason Category column) resolve from
    #      Reference!C:D via VLOOKUP. Previously those were hand-typed
    #      categories that drifted from the canonical Stage 1 list --
    #      Attendance Issues / Job Abandonment / Performance Issues
    #      were all "Other" here but "Involuntary" in Reference. Making
    #      this a lookup means the category is defined in exactly one
    #      place (Reference!D) and every Dashboard stat derived from
    #      this column updates in lockstep.
    for r in range(56, 69):
        d.cell(r, 2).value = (
            f'=IFERROR(VLOOKUP(A{r},Reference!$C$2:$D$14,2,FALSE),"Other")'
        )
    print("[stage3] Data!B56:B68 (reason Category) now VLOOKUPs Reference!$C:$D")

    # 3b. Vol / Invol / Other aggregation via SUMPRODUCT over the
    #     reason table (A56:G68), filtered by category in column B.
    #     Range widened to include the new row.
    for i, (_, _sheet_name) in enumerate(FY_SHEETS):
        data_row = 2 + i  # rows 2..6
        reason_col = chr(ord("C") + i)  # C..G
        d.cell(data_row, 6).value = (
            f'=SUMPRODUCT((Data!$B$56:$B$68="Voluntary")'
            f'*Data!${reason_col}$56:${reason_col}$68)'
        )
        d.cell(data_row, 7).value = (
            f'=SUMPRODUCT((Data!$B$56:$B$68="Involuntary")'
            f'*Data!${reason_col}$56:${reason_col}$68)'
        )
        d.cell(data_row, 8).value = (
            f'=SUMPRODUCT((Data!$B$56:$B$68="Other")'
            f'*Data!${reason_col}$56:${reason_col}$68)'
        )
    print("[stage3] Data!F2:H6 rewritten as SUMPRODUCT over reason categories")

    # 3c. Avg Monthly Separations uses months-elapsed, not months-with-data.
    #     For each FY: 0 if future, 12 if past, MONTH(TODAY()) if current.
    #     FY year is extracted from the A column text "FY 2023" -> 2023.
    for i in range(5):
        data_row = 2 + i
        d.cell(data_row, 5).value = (
            f"=IFERROR(C{data_row}/IF(VALUE(RIGHT(A{data_row},4))<YEAR(TODAY()),12,"
            f"IF(VALUE(RIGHT(A{data_row},4))>YEAR(TODAY()),NA(),MONTH(TODAY()))),0)"
        )
    print("[stage3] Data!E2:E6 rewritten to use months-elapsed")

    # ------------------------------------------------------------------
    # 4. Dashboard -- new rows in Additional Statistics section
    # ------------------------------------------------------------------
    dash = wb["Dashboard"]
    # Row 46: section header "ADDITIONAL STATISTICS"
    # Row 48: Eligible for Rehire Rate
    # Row 49: Exit Interview Completion Rate
    # Rows 50-52: empty -> available for new stats
    dash.cell(50, 1).value = "Rolling 12-month Turnover:"
    # Separations in last 365 days summed across every FY sheet, divided
    # by current-FY active headcount. Wrapped in IFERROR so a missing
    # headcount doesn't break the cell.
    fy_terms = []
    for _, sheet_name in FY_SHEETS:
        fy_terms.append(
            f'SUMPRODUCT((ISNUMBER(\'{sheet_name}\'!$B$9:$B$412))'
            f'*((\'{sheet_name}\'!$B$9:$B$412>=TODAY()-365)'
            f'*(\'{sheet_name}\'!$B$9:$B$412<=TODAY())))'
        )
    sum_expr = "+".join(fy_terms)
    dash.cell(50, 3).value = (
        f"=IFERROR(({sum_expr})/VLOOKUP($B$5,Data!$A$2:$K$6,2,FALSE),0)"
    )

    dash.cell(51, 1).value = "Avg Tenure at Separation:"
    # FY 2026 is the first year with DoH dates, so avg tenure only
    # works for FY 2026 onwards. INDIRECT lets the formula follow the
    # currently selected FY ($B$5). Returns "N/A" rather than 0 when
    # no tenure data exists.
    sheet_b = (
        'INDIRECT("\'"&$B$5&" (Jan"&RIGHT($B$5,2)&"-Dec"&RIGHT($B$5,2)&")\'!B9:B412")'
    )
    sheet_c = (
        'INDIRECT("\'"&$B$5&" (Jan"&RIGHT($B$5,2)&"-Dec"&RIGHT($B$5,2)&")\'!C9:C412")'
    )
    count_expr = f'SUMPRODUCT(({sheet_b}<>"")*({sheet_c}<>""))'
    sum_expr = (
        f'SUMPRODUCT(({sheet_b}<>"")*({sheet_c}<>"")*({sheet_b}-{sheet_c}))'
    )
    dash.cell(51, 3).value = (
        f'=IFERROR(IF({count_expr}=0,"N/A (no DOH data)",'
        f'({sum_expr})/{count_expr}/365.25),"N/A (no DOH data)")'
    )
    print("[stage3] Dashboard rows 50-51 added (rolling 12mo TOR, avg tenure)")

    wb.save(SRC)
    print(f"[stage3] wrote {SRC}")

    # Re-inject DVs (unchanged from Stage 2 -- still strict)
    strict = lambda last: [
        DVSpec(formula="Reference!$B$2:$B$3", sqref=f"E9:E{last}", style="stop",
               error_title="Invalid Rehire", error="Rehire Eligibility must be Yes or No."),
        DVSpec(formula="Reference!$G$2:$G$3", sqref=f"F9:F{last}", style="stop",
               error_title="Invalid Status", error="Status must be U or D."),
        DVSpec(formula="Reference!$E$2:$E$64", sqref=f"G9:G{last}", style="stop",
               error_title="Invalid Location", error="Pick a location from the Reference list."),
        DVSpec(formula="Reference!$C$2:$C$14", sqref=f"H9:H{last}", style="stop",
               error_title="Invalid Reason", error="Pick a reason from the Reference list."),
        DVSpec(formula="Reference!$A$2:$A$18", sqref=f"K9:K{last}", style="stop",
               error_title="Invalid Job", error="Pick a job from the Reference list."),
        DVSpec(formula="Reference!$L$2:$L$10", sqref=f"M9:M{last}", style="stop",
               error_title="Invalid Department", error="Pick a department from the Reference list."),
    ]
    specs = {name: strict(413) for _, name in FY_SHEETS}
    apply_dvs(SRC, specs)
    print("[stage3] re-applied strict x14 data validations")


if __name__ == "__main__":
    main()
