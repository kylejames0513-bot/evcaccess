"""
Stage 2 -- FY sheet schema standardization + strict validation.

What this fixes:
  1. FY 2026 irregular layout. The other four years use a uniform block:
       month header row, col header row, 30 data rows, subtotal, blank
     (34 rows per month, blocks start at row 7, 41, 75, ... 381). FY 2026
     drifted to a completely different row set (9-17, 22-32, 37-47 ...)
     because months were hand-expanded one at a time. This rebuilds
     FY 2026 in place to match the standard template, relocating every
     existing data row to its new position and rebuilding merge cells,
     subtotal formulas, and Data-sheet cross-references in lockstep.

  2. Hardcoded literal subtotal at FY 2026 row 51 (=10). Replaced with a
     real =COUNTA formula. The accompanying "Was not included... do not
     count below" note from old row 48 is preserved as a cell comment on
     the March header so the context isn't lost.

  3. Data! sheet FY 2026 row references. Previously pointed at the
     irregular rows (B18, B33, B51, B85...). Rewrites to the standard
     rows (B39, B73, B107, B141...).

  4. Table1 (an orphaned Excel Table at the old A36:M47 range covering
     the old March block) -- dropped. Nothing references it.

  5. Safe historical data normalization across every FY sheet:
       - strip leading/trailing whitespace on every string cell
       - col E (rehire): "y"/"yes"/"YES" -> "Yes", "n"/"no"/"N" -> "No"
       - col F (status): "u" -> "U", "d" -> "D"
     Anything ambiguous (e.g. "NA", "Yes - Turned in Notice...") is left
     alone rather than guessed.

  6. Data validation flipped from errorStyle="warning" to "stop" so new
     entries that don't match the Reference lists are refused outright.
     FY 2026 validation sqrefs widened from 9:357 to 9:413 now that its
     layout matches the other years.
"""
from __future__ import annotations

import sys
import zipfile
import shutil
import re
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "FY Separation Summary.xlsx"
BACKUP = ROOT / "FY Separation Summary.backup.xlsx"

sys.path.insert(0, str(Path(__file__).parent))
from dv_fixup import DVSpec, apply_dvs  # noqa: E402


# ------------------------------------------------------------------
# Standard FY sheet layout
# ------------------------------------------------------------------
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
BLOCK_SIZE = 34  # rows per month block
FIRST_MONTH_ROW = 7  # "JANUARY YYYY" header row
COLUMNS = [
    "Name", "Date of Separation", "DOH", "Length of Service",
    "Eligible for Rehire", "Status", "Location", "Reason for Leaving",
    "Supervisor", "Exit Interview Date", "Job", "Comments", "Department",
]


def block_rows(month_idx: int) -> tuple[int, int, int, int]:
    """Returns (month_header, col_header, data_start, data_end) for month
    index 0..11. Subtotal lives at data_end + 1."""
    month_header = FIRST_MONTH_ROW + month_idx * BLOCK_SIZE
    col_header = month_header + 1
    data_start = month_header + 2
    data_end = month_header + 31  # 30 data rows
    return month_header, col_header, data_start, data_end


def los_formula(row: int) -> str:
    """Length-of-service DATEDIF expression, matching the existing FY 2026
    template. Returns "" if DoS or DoH is blank."""
    b, c = f"B{row}", f"C{row}"
    return (
        f'=IF(OR({b}="",{c}=""),"",'
        f'IF(DATEDIF({c},{b},"y")>=1,'
        f'DATEDIF({c},{b},"y")&IF(DATEDIF({c},{b},"y")=1," year "," years ")&'
        f'DATEDIF({c},{b},"ym")&IF(DATEDIF({c},{b},"ym")=1," month"," months"),'
        f'IF(DATEDIF({c},{b},"m")>=1,'
        f'DATEDIF({c},{b},"m")&IF(DATEDIF({c},{b},"m")=1," month"," months"),'
        f'DATEDIF({c},{b},"d")&IF(DATEDIF({c},{b},"d")=1," day"," days"))))'
    )


def dept_formula(row: int) -> str:
    """Department lookup for a given data row. Mirrors what's on every FY
    2026 row today. (Stage 3 replaces the lookup table; the formula is
    still valid against the current I2:J175 mapping.)"""
    return (
        f'=IFERROR(VLOOKUP(G{row},Reference!$I$2:$J$175,2,FALSE()),"Operations")'
    )


# ------------------------------------------------------------------
# FY 2026 rebuild (this is the only year whose layout drifted)
# ------------------------------------------------------------------
OLD_FY26_BLOCKS = {
    # Name -> (data_start, data_end) in the old irregular layout. The old
    # month-header row sits two rows above data_start.
    "January":   (9, 17),
    "February":  (22, 32),
    "March":     (37, 47),  # 48 is the "do not count" note row, handled separately
    "April":     (55, 84),
    "May":       (89, 118),
    "June":      (123, 152),
    "July":      (157, 186),
    "August":    (191, 220),
    "September": (225, 254),
    "October":   (259, 288),
    "November":  (293, 322),
    "December":  (327, 356),
}
FY26_MARCH_DONOTCOUNT_ROW = 48


def rebuild_fy2026(ws) -> dict:
    """Return a dict with summary info. Data is relocated in-place."""
    # 1. Snapshot old data
    snapshot: dict[str, list[list]] = {}
    for month, (start, end) in OLD_FY26_BLOCKS.items():
        rows = []
        for r in range(start, end + 1):
            cells = [ws.cell(r, c).value for c in range(1, 14)]
            # keep rows that have ANY content in the content-bearing cols
            has_content = any(
                v not in (None, "")
                for v in (cells[0], cells[1], cells[4], cells[5], cells[6], cells[7])
            )
            if has_content:
                rows.append(cells)
        snapshot[month] = rows

    march_note = ws.cell(FY26_MARCH_DONOTCOUNT_ROW, 1).value

    # 2. Unmerge every merge that touches rows >= 7 (we'll rebuild them).
    # ws.merged_cells.ranges is a set; take a static list first.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= FIRST_MONTH_ROW:
            ws.unmerge_cells(str(rng))

    # 3. Clear every cell from row 7 to 413 (values + formulas). Leave
    # formatting alone -- openpyxl keeps styles when you reset values.
    for r in range(FIRST_MONTH_ROW, 414):
        for c in range(1, 14):
            ws.cell(r, c).value = None

    # 4. Re-emit the standard template, relocating data
    for i, month in enumerate(MONTHS):
        mh, ch, ds, de = block_rows(i)

        # Month header row + merge
        ws.cell(mh, 1).value = f"{month.upper()} 2026"
        ws.merge_cells(start_row=mh, start_column=1, end_row=mh, end_column=13)

        # Column headers
        for c, label in enumerate(COLUMNS, start=1):
            ws.cell(ch, c).value = label

        # Data rows
        for r_offset, row_vals in enumerate(snapshot.get(month, [])):
            r = ds + r_offset
            if r > de:
                # overflow: the month has more than 30 entries. This
                # shouldn't happen with current data (max March=11) but
                # we bail loudly rather than truncate.
                raise RuntimeError(
                    f"FY 2026 {month}: {len(snapshot[month])} entries > 30-row block"
                )
            for c, v in enumerate(row_vals, start=1):
                if c in (4, 13):
                    # LoS (col D) and Dept (col M) get fresh formulas below
                    continue
                ws.cell(r, c).value = v

        # Formulas on every data row (not just populated ones), matching
        # the existing FY2026 template.
        for r in range(ds, de + 1):
            ws.cell(r, 4).value = los_formula(r)
            ws.cell(r, 13).value = dept_formula(r)

        # Subtotal row
        st_row = de + 1
        ws.cell(st_row, 1).value = "SUBTOTAL:"
        ws.cell(st_row, 2).value = f"=COUNTA(A{ds}:A{de})"

    # 5. Preserve the March "do not count" note as a cell comment on the
    # March month-header, so the historical context isn't dropped.
    if march_note:
        march_mh = block_rows(2)[0]  # March is index 2
        ws.cell(march_mh, 1).comment = Comment(
            "Preserved from historical layout (old row 48):\n\n"
            f'"{march_note}"\n\n'
            "The March subtotal is now a live =COUNTA formula. If you "
            "intentionally want to exclude one entry, add an 'Excluded' "
            "marker column rather than hard-coding the subtotal.",
            "Stage 2 cleanup",
        )

    # 6. Drop the orphan Table1 (was bound to A36:M47 in the old layout)
    if "Table1" in ws.tables:
        del ws.tables["Table1"]

    return {
        "months": {m: len(snapshot.get(m, [])) for m in MONTHS},
        "march_note_preserved": bool(march_note),
    }


# ------------------------------------------------------------------
# Data! sheet FY 2026 row references
# ------------------------------------------------------------------
def fix_data_fy26_refs(wb) -> None:
    """Data!B14:M14 currently references the irregular FY 2026 rows.
    Rewrite to the standard rows."""
    d = wb["Data"]
    standard_rows = [39, 73, 107, 141, 175, 209, 243, 277, 311, 345, 379, 413]
    for i, row_num in enumerate(standard_rows):
        col = 2 + i  # B..M
        d.cell(14, col).value = f"='FY 2026 (Jan26-Dec26)'!B{row_num}"


# ------------------------------------------------------------------
# Safe historical data normalization
# ------------------------------------------------------------------
REHIRE_MAP = {
    "y": "Yes", "yes": "Yes", "YES": "Yes", "Yes": "Yes",
    "n": "No", "no": "No", "NO": "No", "No": "No",
}
STATUS_MAP = {"u": "U", "U": "U", "d": "D", "D": "D"}


def normalize_history(ws, last_row: int) -> dict:
    """Trim whitespace and safe case-fix on every data row 9..last_row."""
    changed = {"trimmed": 0, "rehire": 0, "status": 0}
    for r in range(9, last_row + 1):
        for c in range(1, 14):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            stripped = v.strip()
            # Column-specific remap
            if c == 5:  # Eligible for Rehire
                mapped = REHIRE_MAP.get(stripped)
                if mapped and mapped != v:
                    ws.cell(r, c).value = mapped
                    changed["rehire"] += 1
                    continue
            if c == 6:  # Status
                mapped = STATUS_MAP.get(stripped)
                if mapped and mapped != v:
                    ws.cell(r, c).value = mapped
                    changed["status"] += 1
                    continue
            # Generic trim for every other string cell
            if stripped != v:
                ws.cell(r, c).value = stripped
                changed["trimmed"] += 1
    return changed


# ------------------------------------------------------------------
# Strict data validation specs (reused from stage 1 but tightened)
# ------------------------------------------------------------------
def strict_fy_specs(last_row: int) -> list[DVSpec]:
    return [
        DVSpec(
            formula="Reference!$B$2:$B$3",
            sqref=f"E9:E{last_row}",
            style="stop",
            error_title="Invalid Rehire",
            error="Rehire Eligibility must be Yes or No.",
        ),
        DVSpec(
            formula="Reference!$G$2:$G$3",
            sqref=f"F9:F{last_row}",
            style="stop",
            error_title="Invalid Status",
            error="Status must be U (uneligible) or D (discharged).",
        ),
        DVSpec(
            formula="Reference!$E$2:$E$64",
            sqref=f"G9:G{last_row}",
            style="stop",
            error_title="Invalid Location",
            error="Pick a location from the Reference list.",
        ),
        DVSpec(
            formula="Reference!$C$2:$C$14",
            sqref=f"H9:H{last_row}",
            style="stop",
            error_title="Invalid Reason",
            error="Pick a reason from the Reference list.",
        ),
        DVSpec(
            formula="Reference!$A$2:$A$18",
            sqref=f"K9:K{last_row}",
            style="stop",
            error_title="Invalid Job",
            error="Pick a job from the Reference list.",
        ),
        DVSpec(
            formula="Reference!$L$2:$L$10",
            sqref=f"M9:M{last_row}",
            style="stop",
            error_title="Invalid Department",
            error="Pick a department from the Reference list.",
        ),
    ]


# ------------------------------------------------------------------
# Post-save cleanup: drop orphaned Table1 xml + rel
# ------------------------------------------------------------------
def drop_orphan_table(book_path: Path) -> None:
    """openpyxl's `del ws.tables[...]` removes the table object, but the
    table.xml file in the package may still linger in some combinations.
    This is a belt-and-suspenders pass that surgically deletes xl/tables/table1.xml,
    removes the tablePart rel from sheet4's rel file, and drops the
    <tableParts> element from sheet4.xml itself. Also removes the
    Override content type entry for the table part."""
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=str(book_path.parent)
    ) as tmp:
        tmp_path = Path(tmp.name)

    try:
        with zipfile.ZipFile(book_path) as zin:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    name = item.filename
                    if name == "xl/tables/table1.xml":
                        continue  # drop it
                    data = zin.read(name)
                    if name == "[Content_Types].xml":
                        xml = data.decode("utf-8")
                        xml = re.sub(
                            r'<Override[^/]*PartName="/xl/tables/table1\.xml"[^/]*/>',
                            "",
                            xml,
                        )
                        data = xml.encode("utf-8")
                    elif name == "xl/worksheets/_rels/sheet4.xml.rels":
                        xml = data.decode("utf-8")
                        xml = re.sub(
                            r'<Relationship\b[^>]*Target="[^"]*tables/table1\.xml"[^>]*>',
                            "",
                            xml,
                        )
                        data = xml.encode("utf-8")
                    elif name == "xl/worksheets/sheet4.xml":
                        xml = data.decode("utf-8")
                        xml = re.sub(
                            r"<tableParts[^/]*/?>.*?</tableParts>|<tableParts[^/]*/>",
                            "",
                            xml,
                            flags=re.DOTALL,
                        )
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)
        shutil.move(str(tmp_path), str(book_path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"missing {SRC}")
    if not BACKUP.exists():
        raise SystemExit(f"missing backup at {BACKUP}")

    wb = openpyxl.load_workbook(SRC)

    # Phase A: rebuild FY 2026
    fy26 = wb["FY 2026 (Jan26-Dec26)"]
    info = rebuild_fy2026(fy26)
    print(f"[stage2] FY 2026 rebuilt. entries per month: {info['months']}")
    if info["march_note_preserved"]:
        print("[stage2] preserved March 'do not count' note as cell comment")

    # Phase B: fix Data! FY 2026 cross-references
    fix_data_fy26_refs(wb)
    print("[stage2] Data!B14:M14 rewritten to standard FY 2026 rows")

    # Phase C: normalize historical data
    norm_totals = {"trimmed": 0, "rehire": 0, "status": 0}
    for sheet_name, last_row in [
        ("FY 2023 (Jan23-Dec23)", 413),
        ("FY 2024 (Jan24-Dec24)", 413),
        ("FY 2025 (Jan25-Dec25)", 413),
        ("FY 2026 (Jan26-Dec26)", 413),
        ("FY 2027 (Jan27-Dec27)", 413),
    ]:
        ws = wb[sheet_name]
        counts = normalize_history(ws, last_row)
        for k, v in counts.items():
            norm_totals[k] += v
    print(f"[stage2] normalized historical data: {norm_totals}")

    wb.save(SRC)
    print(f"[stage2] wrote {SRC}")

    # Phase D: belt-and-suspenders drop of orphan Table1
    drop_orphan_table(SRC)
    print("[stage2] dropped orphan Table1 (old March block)")

    # Phase E: strict data validations
    specs = {
        "FY 2023 (Jan23-Dec23)": strict_fy_specs(413),
        "FY 2024 (Jan24-Dec24)": strict_fy_specs(413),
        "FY 2025 (Jan25-Dec25)": strict_fy_specs(413),
        "FY 2026 (Jan26-Dec26)": strict_fy_specs(413),
        "FY 2027 (Jan27-Dec27)": strict_fy_specs(413),
    }
    apply_dvs(SRC, specs)
    print("[stage2] re-applied x14 data validations with errorStyle=stop")


if __name__ == "__main__":
    main()
