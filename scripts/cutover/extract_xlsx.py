#!/usr/bin/env python3
"""
Cutover Stage 5: extract historical data from the EVC xlsx workbook
into JSONL files under /tmp/evc_cutover/. The intermediate JSONL files
are then driven into Supabase via SQL bulk inserts (see ingest_*.py).

Each output file is one JSON object per line. Dates are normalized to
ISO YYYY-MM-DD strings. Empty cells are dropped at this stage.
"""
import json
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl

XLSX_PATH = Path("Google Sheets/EVC_Attendance_Tracker (2).xlsx")
OUT_DIR = Path("/tmp/evc_cutover")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value).strip()
    if not s:
        return None
    # ISO already
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # US m/d/y
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        mo, d, y = m.groups()
        if len(y) == 2:
            yi = int(y)
            y = f"19{y}" if yi >= 70 else f"20{y.zfill(2)}"
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return None


def write_jsonl(name, rows):
    out = OUT_DIR / f"{name}.jsonl"
    with out.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, default=str))
            fh.write("\n")
    print(f"  wrote {out} ({len(rows)} rows)")


def header_row(ws, row_num):
    return [
        (ws.cell(row=row_num, column=c).value or "")
        if isinstance(ws.cell(row=row_num, column=c).value, str)
        else ws.cell(row=row_num, column=c).value
        for c in range(1, ws.max_column + 1)
    ]


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=False)

    # ────────────────────────────────────────
    # Employees
    # ────────────────────────────────────────
    print("Extracting Employees tab...")
    ws = wb["Employees"]
    headers = header_row(ws, 1)
    employees = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h}
        last = (row.get("Last Name") or "").strip() if isinstance(row.get("Last Name"), str) else row.get("Last Name")
        first = (row.get("First Name") or "").strip() if isinstance(row.get("First Name"), str) else row.get("First Name")
        if not last or not first:
            continue
        employees.append(
            {
                "last_name": str(last).strip(),
                "first_name": str(first).strip(),
                "preferred_name": (str(row.get("Preferred Name")).strip() if row.get("Preferred Name") else None),
                "paylocity_id": (str(row.get("ID")).strip() if row.get("ID") else None),
                "job_title": (str(row.get("Position / Job Title")).strip() if row.get("Position / Job Title") else None),
                "position": (str(row.get("Position / Job Title")).strip() if row.get("Position / Job Title") else None),
                "hire_date": to_iso_date(row.get("Hire Date")),
                "division": (str(row.get("Division")).strip() if row.get("Division") else None),
                "department": (str(row.get("Department")).strip() if row.get("Department") else None),
                "status": (str(row.get("Status")).strip() if row.get("Status") else None),
            }
        )
    write_jsonl("employees", employees)

    # ────────────────────────────────────────
    # Access (wide -> long)
    # ────────────────────────────────────────
    print("Extracting Access tab (wide -> long)...")
    ws = wb["Access"]
    headers = header_row(ws, 1)
    completions = []
    excusals = []
    skip_columns = {"L NAME", "F NAME", "ACTIVE"}
    for r in range(2, ws.max_row + 1):
        last = ws.cell(row=r, column=1).value
        first = ws.cell(row=r, column=2).value
        active = ws.cell(row=r, column=3).value
        if not last or not first:
            continue
        last = str(last).strip()
        first = str(first).strip()
        for ci, hdr in enumerate(headers):
            if not hdr or hdr in skip_columns:
                continue
            cell = ws.cell(row=r, column=ci + 1).value
            if cell is None:
                continue
            iso = to_iso_date(cell)
            if iso:
                completions.append(
                    {
                        "last_name": last,
                        "first_name": first,
                        "column_key": hdr,
                        "completion_date": iso,
                    }
                )
            else:
                s = str(cell).strip()
                if s:
                    excusals.append(
                        {
                            "last_name": last,
                            "first_name": first,
                            "column_key": hdr,
                            "reason": s,
                        }
                    )
    write_jsonl("access_completions", completions)
    write_jsonl("access_excusals", excusals)

    # ────────────────────────────────────────
    # Paylocity Import
    # ────────────────────────────────────────
    print("Extracting Paylocity Import tab...")
    ws = wb["Paylocity Import"]
    headers = header_row(ws, 1)
    skip_codes = {"DL", "MVR", "Insurance", "Background", "Veh Ins Declination", "Driver's License", "Vehicle Insurance Declination Page"}
    paylocity = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h}
        eid = row.get("Employee Id")
        if eid is None:
            continue
        eid = str(eid).strip()
        if not eid:
            continue
        skill = (str(row.get("Skill") or "")).strip()
        code = (str(row.get("Code") or "")).strip()
        raw = code if code else skill
        if raw in skip_codes:
            continue
        eff = to_iso_date(row.get("Effective/Issue Date"))
        exp = to_iso_date(row.get("Expiration Date"))
        if not eff:
            continue
        paylocity.append(
            {
                "paylocity_id": eid,
                "raw_training": raw,
                "skill": skill,
                "code": code,
                "completion_date": eff,
                "expiration_date": exp,
                "last_name": (str(row.get("Last Name")) if row.get("Last Name") else None),
                "first_name": (str(row.get("First Name")) if row.get("First Name") else None),
            }
        )
    write_jsonl("paylocity", paylocity)

    # ────────────────────────────────────────
    # PHS Import
    # ────────────────────────────────────────
    print("Extracting PHS Import tab...")
    ws = wb["PHS Import"]
    headers = header_row(ws, 1)
    phs = []
    phs_special = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h}
        name = row.get("Employee Name")
        if not name:
            continue
        cat = (str(row.get("Upload Category") or "")).strip()
        typ = (str(row.get("Upload Type") or "")).strip()

        if cat == "Drivers License":
            continue

        if cat == "Med Admin" and typ in ("No Show", "Fail"):
            phs_special.append(
                {
                    "full_name": str(name).strip(),
                    "category": cat,
                    "type": typ,
                    "effective_date": to_iso_date(row.get("Effective Date")),
                }
            )
            continue

        # Map (cat, type) -> raw_training name
        if cat == "Med Admin" and typ == "Certification":
            raw = "Med Recert"
        elif cat == "CPR/FA":
            raw = "CPR/FA"
        elif cat == "Additional Training":
            raw = typ
        else:
            raw = typ if typ else cat

        eff = to_iso_date(row.get("Effective Date"))
        exp = to_iso_date(row.get("Expiration Date"))
        if not eff:
            continue
        phs.append(
            {
                "full_name": str(name).strip(),
                "raw_training": raw,
                "category": cat,
                "type": typ,
                "completion_date": eff,
                "expiration_date": exp,
            }
        )
    write_jsonl("phs", phs)
    write_jsonl("phs_special", phs_special)

    # ────────────────────────────────────────
    # Training Records (legacy sign-in)
    # ────────────────────────────────────────
    print("Extracting Training Records tab...")
    ws = wb["Training Records"]
    headers = header_row(ws, 1)
    signin = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h}
        attendee = row.get("Attendee Name")
        session = row.get("Training Session")
        date = row.get("Date of Training")
        if not attendee or not session:
            continue
        iso = to_iso_date(date)
        if not iso:
            continue
        signin.append(
            {
                "attendee_name": str(attendee).strip(),
                "raw_training": str(session).strip(),
                "completion_date": iso,
                "pass_fail": (str(row.get("Pass / Fail")).strip() if row.get("Pass / Fail") else None),
                "reviewed_by": (str(row.get("Reviewed By")).strip() if row.get("Reviewed By") else None),
            }
        )
    write_jsonl("signin", signin)

    print("Done.")


if __name__ == "__main__":
    main()
